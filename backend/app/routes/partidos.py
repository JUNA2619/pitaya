from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from app.database import supabase
from app.routes.auth import verificar_token
import csv
import io
import openpyxl

router = APIRouter()

@router.get("")
def listar_partidos(usuario=Depends(verificar_token)):
    resultado = supabase.table("partidos").select("*").order("fecha").order("hora").execute()
    return resultado.data

@router.post("")
def crear_partido(data: dict, usuario=Depends(verificar_token)):
    partido = {
        "torneo": str(data.get("torneo", "") or "").strip() or "Sin torneo",
        "cancha": str(data.get("cancha", "") or "").strip() or "Sin cancha",
        "fecha": data.get("fecha", ""),
        "hora": data.get("hora", ""),
        "tipo": str(data.get("tipo", "") or "futbol").strip(),
        "num_periodos": int(data.get("num_periodos") or 2),
        "tiempo_periodo": int(data.get("tiempo_periodo") or 20),
        "tipo_pago": str(data.get("tipo_pago", "") or "en_cancha").strip(),
        "equipos": str(data.get("equipos", "") or "").strip() or None,
        "estado": "sin_asignar"
    }
    if not partido["fecha"] or not partido["hora"]:
        raise HTTPException(status_code=400, detail="La fecha y la hora del partido son obligatorias.")
    resultado = supabase.table("partidos").insert(partido).execute()
    return resultado.data[0]

@router.delete("/{partido_id}")
def eliminar_partido(partido_id: str, usuario=Depends(verificar_token)):
    supabase.table("partidos").delete().eq("id", partido_id).execute()
    return {"ok": True}

def normalizar_tipo(valor):
    if not valor:
        return "futbol"
    v = str(valor).strip().lower().replace("á", "a").replace("é", "e")
    if "sala" in v:
        return "futbol_sala"
    return "futbol"

def normalizar_pago(valor):
    if not valor:
        return "en_cancha"
    v = str(valor).strip().lower().replace("_", " ")
    if "pendiente" in v:
        return "pendiente"
    return "en_cancha"

def parsear_filas(filas):
    partidos = []
    errores = []
    for i, fila in enumerate(filas, start=2):
        fecha = str(fila.get("fecha", "") or "").strip()
        hora = str(fila.get("hora", "") or "").strip()
        if not fecha and not hora:
            continue
        if not fecha or not hora:
            errores.append(f"Fila {i}: falta {'fecha' if not fecha else 'hora'}")
            continue
        try:
            partidos.append({
                "torneo": str(fila.get("torneo", "") or "").strip() or "Sin torneo",
                "cancha": str(fila.get("cancha", "") or "").strip() or "Sin cancha",
                "fecha": fecha,
                "hora": hora,
                "tipo": normalizar_tipo(fila.get("tipo")),
                "num_periodos": int(fila.get("periodos") or fila.get("num_periodos") or 2),
                "tiempo_periodo": int(fila.get("minutos") or fila.get("tiempo_periodo") or 20),
                "tipo_pago": normalizar_pago(fila.get("pago") or fila.get("tipo_pago")),
                "equipos": str(fila.get("equipos", "") or "").strip() or None,
                "estado": "sin_asignar"
            })
        except Exception as e:
            errores.append(f"Fila {i}: error — {str(e)}")
    return partidos, errores

@router.post("/excel")
async def subir_excel(file: UploadFile = File(...), usuario=Depends(verificar_token)):
    nombre = file.filename.lower()
    if not nombre.endswith(".xlsx") and not nombre.endswith(".xls"):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsx o .xls")

    contenido = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        filas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None and str(v).strip() != "" for v in row):
                filas.append(dict(zip(headers, row)))
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo. Verifique que sea un Excel válido.")

    partidos, errores = parsear_filas(filas)

    if not partidos:
        raise HTTPException(status_code=400, detail=f"No se importó ningún partido. Verifique que las columnas fecha y hora estén diligenciadas. Errores: {errores}")

    resultado = supabase.table("partidos").insert(partidos).execute()
    return {
        "importados": len(resultado.data),
        "errores": errores
    }